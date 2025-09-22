import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime

current_dir = os.path.dirname(os.path.abspath(__file__))
input_file = os.path.join(current_dir, '../temp_files/按日期分表的处理打卡数据.xlsx')
output_file = os.path.join(current_dir, '../temp_files/按打卡时间分列的打卡数据.xlsx')

# 加载工作簿
wb = openpyxl.load_workbook(input_file)
# 创建新工作簿用于保存处理后的数据
new_wb = openpyxl.Workbook()
# 移除默认创建的工作表
default_sheet = new_wb.active
new_wb.remove(default_sheet)

# 复制"原始数据"工作表到新工作簿
if '原始数据' in wb.sheetnames:
    original_sheet = wb['原始数据']
    new_ws = new_wb.create_sheet(title='原始数据')
    for row in original_sheet.iter_rows(values_only=True):
        new_ws.append(row)

# 处理其他工作表
processed_sheets = 0
for sheet_name in wb.sheetnames:
    # 跳过"原始数据"工作表
    if sheet_name == '原始数据':
        continue

    # 读取当前工作表数据
    df = pd.read_excel(input_file, sheet_name=sheet_name)

    # 检查是否包含"打卡时间"列
    if '打卡时间' not in df.columns:
        print(f"警告：工作表 '{sheet_name}' 中未找到 '打卡时间' 列，已跳过")
        continue

    # 处理打卡时间列，按";"拆分
    # 最多拆分为4列（第一次到第四次打卡）
    punch_times = df['打卡时间'].str.split(';', expand=True, n=3)
    punch_times.columns = ['第一次打卡', '第二次打卡', '第三次打卡', '第四次打卡']

    # 合并原始数据和拆分后的打卡时间
    # 先删除原始的"打卡时间"列，再合并
    result_df = df.drop(columns=['打卡时间']).join(punch_times)

    # 确保时间格式正确（去除可能的空字符）
    for col in ['第一次打卡', '第二次打卡', '第三次打卡', '第四次打卡']:
        if col in result_df.columns:
            result_df[col] = result_df[col].str.strip()


    # 新增：添加班次列
    def determine_shift(row):
        # 后勤部班次为空
        if row.get('部门') == '后勤部':
            return ''

        first_punch = row.get('第一次打卡')
        if not first_punch:
            return ''

        try:
            # 解析时间
            punch_time = datetime.strptime(first_punch, '%H:%M').time()
            # 定义时间界限
            noon = datetime.strptime('12:00', '%H:%M').time()
            five_pm = datetime.strptime('17:00', '%H:%M').time()
            ten_pm = datetime.strptime('23:59', '%H:%M').time()

            # 判断班次
            if punch_time < noon:
                return '早班'
            elif noon <= punch_time < five_pm:
                return '中班'
            elif five_pm <= punch_time < ten_pm:
                return '晚班'
            else:
                return ''
        except:
            # 时间格式错误时返回空
            return ''


    # 应用函数计算班次
    result_df['班次'] = result_df.apply(determine_shift, axis=1)

    # 调整列顺序：将班次列移到第一次打卡列前面
    cols = result_df.columns.tolist()
    # 找到第一次打卡的索引位置
    first_punch_idx = cols.index('第一次打卡')
    # 移除班次列并插入到第一次打卡前面
    cols.remove('班次')
    cols.insert(first_punch_idx, '班次')
    result_df = result_df[cols]

    # 创建新工作表并写入处理后的数据
    new_ws = new_wb.create_sheet(title=sheet_name)
    for r in dataframe_to_rows(result_df, index=False, header=True):
        new_ws.append(r)

    processed_sheets += 1
    print(f"已处理工作表：{sheet_name}，拆分了 {len(result_df)} 条打卡记录")

# 保存处理后的文件
new_wb.save(output_file)
print(f"\n处理完成！共处理 {processed_sheets} 个工作表，结果已保存到 {output_file}")