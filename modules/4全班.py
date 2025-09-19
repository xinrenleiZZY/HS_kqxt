import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timedelta
import os

current_dir = os.path.dirname(os.path.abspath(__file__))
input_file = os.path.join(current_dir, '../temp_files/按打卡时间分列的打卡数据.xlsx')
output_file = os.path.join(current_dir, '../temp_files/全班次处理后的打卡数据.xlsx')


# 系统休息时间：次日05:00
SYSTEM_REST_TIME = datetime.strptime("05:00", "%H:%M").time()


# 时间处理辅助函数
def parse_time(time_str):
    """将时间字符串转换为datetime.time对象，失败返回None"""
    if pd.isna(time_str) or str(time_str).strip() == "":
        return None
    try:
        cleaned = str(time_str).strip().replace("次日", "").replace("凌晨", "").strip()
        return datetime.strptime(cleaned, "%H:%M").time()
    except:
        return None


def round_down_to_hour(time_obj):
    """将时间向下取整到整点（如20:31→20:00，20:30→20:30）"""
    if time_obj.minute >= 30:
        return time_obj.replace(minute=30, second=0)
    else:
        return time_obj.replace(minute=0, second=0)


def time_diff_in_hours(time1, time2):
    """计算两个时间差（小时），time1 > time2 时返回正值"""
    if not time1 or not time2:
        return 0, 0

    # 处理跨天情况（如23:00到01:00）
    if time1 >= time2:
        diff = timedelta(hours=time1.hour, minutes=time1.minute) - timedelta(hours=time2.hour, minutes=time2.minute)
    else:
        diff = (timedelta(days=1) - timedelta(hours=time2.hour, minutes=time2.minute)
                + timedelta(hours=time1.hour, minutes=time1.minute))

    total_minutes = int(diff.total_seconds() // 60)
    hours = total_minutes // 60
    minutes = total_minutes % 60
    return hours, minutes


def format_time_diff(hours, minutes):
    """将时间差格式化为"X小时Y分钟"字符串"""
    if hours == 0 and minutes == 0:
        return "0分钟"
    parts = []
    if hours > 0:
        parts.append(f"{hours}小时")
    if minutes > 0:
        parts.append(f"{minutes}分钟")
    return "".join(parts)


def process_morning_shift(row):
    """处理早班打卡数据"""
    result = {
        '上班卡类型': "", '迟到时间': "0分钟",
        '中午下班卡类型': "", '中午上班卡类型': "", '白天加班时长(小时)': "",
        '下班卡类型': "", '晚上加班时长(小时)': "", '打卡状态': ""
    }

    # 解析所有打卡时间
    punches = {
        'first': parse_time(row['第一次打卡']),
        'second': parse_time(row['第二次打卡']),
        'third': parse_time(row['第三次打卡']),
        'fourth': parse_time(row['第四次打卡'])
    }

    # 1. 处理上班打卡（早上8:00）
    work_start = datetime.strptime("08:00", "%H:%M").time()
    work_end_am = datetime.strptime("12:00", "%H:%M").time()

    # 确定有效的上班卡
    valid_morning_punch = None
    for p in ['first', 'second']:
        if punches[p]:
            valid_morning_punch = punches[p]
            break

    if valid_morning_punch:
        if valid_morning_punch <= work_start:
            result['上班卡类型'] = "8:00上班卡"
        elif work_start < valid_morning_punch <= work_end_am:
            result['上班卡类型'] = "迟到"
            late_h, late_m = time_diff_in_hours(valid_morning_punch, work_start)
            result['迟到时间'] = format_time_diff(late_h, late_m)
        else:
            result['上班卡类型'] = "缺勤"
    else:
        result['上班卡类型'] = "缺勤"

    # 2. 处理中午打卡（12:00-13:30）
    noon_start = datetime.strptime("12:00", "%H:%M").time()
    noon_end = datetime.strptime("13:30", "%H:%M").time()
    noon_early_end = datetime.strptime("12:30", "%H:%M").time()

    noon_punches = []
    for p in ['second', 'third']:
        if punches[p] and noon_start <= punches[p] <= noon_end:
            noon_punches.append(punches[p])

    if len(noon_punches) == 1:
        result['中午下班卡类型'] = "12:00下班卡"
        result['中午上班卡类型'] = "未打卡"
    elif len(noon_punches) >= 2:
        result['中午下班卡类型'] = "12:00下班卡"
        if noon_punches[1] <= noon_early_end:
            result['中午上班卡类型'] = "12:30上班卡"
            result['白天加班时长(小时)'] = 1
        else:
            result['中午上班卡类型'] = "13:30上班卡"

    # 3. 处理下午下班打卡（17:30）
    work_end_pm = datetime.strptime("17:30", "%H:%M").time()
    eighteen_oclock = datetime.strptime("18:00", "%H:%M").time()  # 18点判断基准

    valid_evening_punch = None
    for p in ['third', 'fourth']:
        if punches[p]:
            valid_evening_punch = punches[p]

    if valid_evening_punch:
        if (valid_evening_punch >= work_end_pm) or (valid_evening_punch <= SYSTEM_REST_TIME):
            rounded_time = round_down_to_hour(valid_evening_punch)
            result['下班卡类型'] = f"{rounded_time.strftime('%H:%M')}下班卡"

            if valid_evening_punch >= work_end_pm:
                ot_h, ot_m = time_diff_in_hours(rounded_time, work_end_pm)
                total_overtime = ot_h + ot_m / 60
                # 晚于18点下班，加班时长减0.5小时
                if valid_evening_punch > eighteen_oclock:
                    total_overtime -= 0.5
            else:
                hours_pm = 24 - (work_end_pm.hour + work_end_pm.minute / 60)
                hours_am = rounded_time.hour + rounded_time.minute / 60
                total_overtime = hours_pm + hours_am
                ot_h = int(total_overtime)
                ot_m = int(round((total_overtime - ot_h) * 60))

            result['晚上加班时长(小时)'] = max(0, round(total_overtime, 1))
        else:
            result['下班卡类型'] = "17:30下班卡-早退"
    else:
        result['下班卡类型'] = "缺卡"

    # 4. 综合打卡状态
    if result['上班卡类型'] == "缺勤":
        result['打卡状态'] = "缺勤"
    elif result['下班卡类型'] == "缺卡":
        result['打卡状态'] = "下班缺卡"
    else:
        result['打卡状态'] = "正常"

    return result


def process_afternoon_shift(row):
    """处理中班打卡数据"""
    result = {
        '上班卡类型': "", '迟到时间': "0分钟",
        '中午下班卡类型': "", '中午上班卡类型': "", '白天加班时长(小时)': "",
        '下班卡类型': "", '晚上加班时长(小时)': "", '打卡状态': ""
    }

    # 解析所有打卡时间
    punches = {
        'first': parse_time(row['第一次打卡']),
        'second': parse_time(row['第二次打卡']),
        'third': parse_time(row['第三次打卡']),
        'fourth': parse_time(row['第四次打卡'])
    }

    # 1. 处理上班打卡（中午13:30）
    work_start = datetime.strptime("13:30", "%H:%M").time()
    work_end_am = datetime.strptime("17:30", "%H:%M").time()

    # 确定有效的上班卡
    valid_afternoon_punch = None
    for p in ['first', 'second']:
        if punches[p]:
            valid_afternoon_punch = punches[p]
            break

    if valid_afternoon_punch:
        if valid_afternoon_punch <= work_start:
            result['上班卡类型'] = "13:30上班卡"
        elif work_start < valid_afternoon_punch <= work_end_am:
            result['上班卡类型'] = "迟到"
            late_h, late_m = time_diff_in_hours(valid_afternoon_punch, work_start)
            result['迟到时间'] = format_time_diff(late_h, late_m)
        else:
            result['上班卡类型'] = "缺勤"
    else:
        result['上班卡类型'] = "缺勤"

    # 2. 处理傍晚打卡（17:30-18:00）
    evening_start = datetime.strptime("17:30", "%H:%M").time()
    evening_end = datetime.strptime("18:00", "%H:%M").time()

    evening_punches = []
    for p in ['second', 'third']:
        if punches[p] and evening_start <= punches[p] <= evening_end:
            evening_punches.append(punches[p])

    if len(evening_punches) == 1:
        result['中午下班卡类型'] = "17:30下班卡"
        result['中午上班卡类型'] = "未打卡"
    elif len(evening_punches) >= 2:
        result['中午下班卡类型'] = "17:30下班卡"
        result['中午上班卡类型'] = "18:00上班卡"

    # 3. 处理晚上下班打卡（22:00）
    work_end_pm = datetime.strptime("22:00", "%H:%M").time()
    system_rest_time = datetime.strptime("07:00", "%H:%M").time() # 系统休息时间：次日07:00

    valid_night_punch = None
    for p in ['fourth']:
        if punches[p]:
            valid_night_punch = punches[p]
            

    if valid_night_punch:
        # 先判断是否为次日打卡（0:00-7:00）
        if valid_night_punch <= system_rest_time:
            rounded_time = round_down_to_hour(valid_night_punch)
            result['下班卡类型'] = f"{rounded_time.strftime('%H:%M')}下班卡"
            # 计算跨天加班时长（22:00到次日打卡时间）
            hours_pm = 24 - (work_end_pm.hour + work_end_pm.minute / 60)
            hours_am = rounded_time.hour + rounded_time.minute / 60
            total_overtime = hours_pm + hours_am
            ot_h = int(total_overtime)
            ot_m = int(round((total_overtime - ot_h) * 60))
            result['晚上加班时长(小时)'] = max(0, round(ot_h + ot_m / 60, 1))
        
        # 再判断是否为当天22:00后打卡
        elif valid_night_punch >= work_end_pm:
            rounded_time = round_down_to_hour(valid_night_punch)
            result['下班卡类型'] = f"{rounded_time.strftime('%H:%M')}下班卡"
            ot_h, ot_m = time_diff_in_hours(rounded_time, work_end_pm)
            result['晚上加班时长(小时)'] = max(0, round(ot_h + ot_m / 60, 1))
        
        # 当天22:00前正常打卡
        elif valid_night_punch <= work_end_pm:
            result['下班卡类型'] = "22:00下班卡-早退"
            result['晚上加班时长(小时)'] = 0.0
        
        # 其他异常时间（7:00-22:00之间非加班时段）
        else:
            result['下班卡类型'] = f"{valid_night_punch.strftime('%H:%M')}（异常下班）"
            result['晚上加班时长(小时)'] = "异常"
    else:
        
        result['下班卡类型'] = "缺卡"

    # 4. 综合打卡状态
    if result['上班卡类型'] == "缺勤":
        result['打卡状态'] = "缺勤"
    elif result['下班卡类型'] == "缺卡":
        result['打卡状态'] = "缺卡"
    elif result['下班卡类型'] == "异常":
        result['打卡状态'] = "异常"
    else:
        result['打卡状态'] = "正常"

    return result


def process_night_shift(row):
    """处理晚班打卡数据"""
    result = {
        '上班卡类型': "", '迟到时间': "0分钟",
        '中午下班卡类型': "", '中午上班卡类型': "", '白天加班时长(小时)': "",
        '下班卡类型': "", '晚上加班时长(小时)': "", '打卡状态': ""
    }

    # 解析所有打卡时间
    punches = {
        'first': parse_time(row['第一次打卡']),
        
        'second': parse_time(row['第二次打卡']),
        'third': None,
        'fourth': None
    }


    # 1. 处理上班打卡（晚上18:00）
    work_start = datetime.strptime("18:00", "%H:%M").time()
    work_end_limit = datetime.strptime("23:00", "%H:%M").time() # 迟到截止时间
    absence_limit = datetime.strptime("02:00", "%H:%M").time()  # 缺勤判定时间
    absence_limit_SYSTEM_REST_TIME = datetime.strptime("09:00", "%H:%M").time()  # 次日9:00

    # 确定有效的上班卡
    valid_night_punch = None
    for p in ['first']:
        if punches[p]:
            valid_night_punch = punches[p]
            break

    if valid_night_punch:
        if valid_night_punch <= work_start:
            result['上班卡类型'] = "18:00上班卡"
        elif work_start < valid_night_punch <= work_end_limit:
            result['上班卡类型'] = "迟到"
            late_h, late_m = time_diff_in_hours(valid_night_punch, work_start)
            result['迟到时间'] = format_time_diff(late_h, late_m)
        else:
            result['上班卡类型'] = "缺勤"
    else:
        # 检查是否在18:00-次日2:00未打卡
        result['上班卡类型'] = "缺勤"

    # 2. 处理次日下班打卡（2:00）
    work_end = datetime.strptime("02:00", "%H:%M").time()
    overtime_limit = datetime.strptime("09:00", "%H:%M").time()  # 加班截止时间

    valid_morning_punch = None
    for p in ['second']:
        if punches[p]:
            valid_morning_punch = punches[p]
            print(valid_morning_punch)

    if valid_morning_punch:
        if valid_morning_punch <= work_end:
            result['下班卡类型'] = "02:00下班卡"
            result['晚上加班时长(小时)'] = 0.0
            # 02:00后且9:00前打卡 - 计算加班
        elif work_end < valid_morning_punch <= overtime_limit:
            rounded_time = round_down_to_hour(valid_morning_punch)
            result['下班卡类型'] = f"{rounded_time.strftime('%H:%M')}下班卡"

            ot_h, ot_m = time_diff_in_hours(rounded_time, work_end)
            result['晚上加班时长(小时)'] = max(0, round(ot_h + ot_m / 60, 1))
    else:
        result['下班卡类型'] = "缺卡"

    #     elif (valid_morning_punch >= work_end) or (valid_morning_punch <= absence_limit_SYSTEM_REST_TIME):
    #         rounded_time = round_down_to_hour(valid_morning_punch)
    #         result['下班卡类型'] = f"{rounded_time.strftime('%H:%M')}下班卡"

    #         if valid_morning_punch >= work_end:
    #             ot_h, ot_m = time_diff_in_hours(rounded_time, work_end)
    #         else:
    #             hours_pm = 24 - (work_end.hour + work_end.minute / 60)
    #             hours_am = rounded_time.hour + rounded_time.minute / 60
    #             total_overtime = hours_pm + hours_am
    #             ot_h = int(total_overtime)
    #             ot_m = int(round((total_overtime - ot_h) * 60))

    #         result['晚上加班时长(小时)'] = max(0, round(ot_h + ot_m / 60, 1))
    #     else:
    #         result['下班卡类型'] = "02:00下班卡"
    # else:
    #     result['下班卡类型'] = "缺卡"

    # 3. 综合打卡状态
    if result['上班卡类型'] == "缺勤":
        result['打卡状态'] = "缺勤"
    elif result['下班卡类型'] == "缺卡":
        result['打卡状态'] = "缺卡"
    else:
        result['打卡状态'] = "正常"

    return result


def process_logistics(row):
    """处理后勤部打卡数据"""
    result = {
        '上班卡类型': "", '迟到时间': "",
        '中午下班卡类型': "", '中午上班卡类型': "", '白天加班时长(小时)': "",
        '下班卡类型': "", '晚上加班时长(小时)': "", '打卡状态': ""
    }

    # 检查是否有任何打卡记录
    has_punch = False
    for col in ['第一次打卡', '第二次打卡', '第三次打卡', '第四次打卡']:
        if pd.notna(row[col]) and str(row[col]).strip() != "":
            has_punch = True
            break

    if has_punch:
        result['打卡状态'] = "正常"
        result['上班卡类型'] = "正常打卡"
    else:
        result['打卡状态'] = "缺勤"
        result['上班卡类型'] = "未打卡"

    return result


# 加载工作簿
wb = openpyxl.load_workbook(input_file)
# 创建新工作簿用于保存处理后的数据
new_wb = openpyxl.Workbook()
# 移除默认创建的工作表
default_sheet = new_wb.active
new_wb.remove(default_sheet)

# 复制"原始数据"工作表到新工作簿
if '原始数据' in wb.sheetnames:
    original_sheet = wb['原始数据']
    new_ws = new_wb.create_sheet(title='原始数据')
    for row in original_sheet.iter_rows(values_only=True):
        new_ws.append(row)

# 处理其他工作表
processed_sheets = 0
for sheet_name in wb.sheetnames:
    if sheet_name == '原始数据':
        continue

    # 读取当前工作表数据
    df = pd.read_excel(input_file, sheet_name=sheet_name)

    # 检查必要的列是否存在
    required_cols = ['姓名', '员工ID', '部门', '班次',
                     '第一次打卡', '第二次打卡', '第三次打卡', '第四次打卡']
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        print(f"警告：工作表 '{sheet_name}' 缺少必要列 {missing_cols}，已跳过")
        continue

    # 新增处理结果列
    result_cols = [
        '上班卡类型', '迟到时间',
        '中午下班卡类型', '中午上班卡类型', '白天加班时长(小时)',
        '下班卡类型', '晚上加班时长(小时)', '打卡状态'
    ]
    for col in result_cols:
        df[col] = ""

    # 按班次处理每一行数据
    for idx, row in df.iterrows():
        department = str(row['部门']).strip()
        shift = str(row['班次']).strip()

        # 后勤部单独处理，不考虑班次
        if department == '后勤部':
            results = process_logistics(row)
        else:
            # 按班次处理
            if shift == '早班':
                results = process_morning_shift(row)
            elif shift == '中班':
                results = process_afternoon_shift(row)
            elif shift == '晚班':
                results = process_night_shift(row)
            else:
                # 未知班次
                results = {col: "未知班次" for col in result_cols}

        # 将处理结果写入DataFrame
        for col, value in results.items():
            df.at[idx, col] = value

    # 创建新工作表并写入处理后的数据
    new_ws = new_wb.create_sheet(title=sheet_name)
    for r in dataframe_to_rows(df, index=False, header=True):
        new_ws.append(r)

    processed_sheets += 1
    print(f"已处理工作表：{sheet_name}，共处理 {len(df)} 条记录")

# 保存处理后的文件
new_wb.save(output_file)
print(f"\n处理完成！共处理 {processed_sheets} 个工作表，结果已保存到 {output_file}")
